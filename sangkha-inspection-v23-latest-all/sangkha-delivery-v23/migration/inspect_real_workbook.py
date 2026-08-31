from openpyxl import load_workbook

path = "/tmp/real_data.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames)
for ws in wb.worksheets:
    print(f"--- {ws.title} rows={ws.max_row} cols={ws.max_column} ---")
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        print(list(row)[:20])
