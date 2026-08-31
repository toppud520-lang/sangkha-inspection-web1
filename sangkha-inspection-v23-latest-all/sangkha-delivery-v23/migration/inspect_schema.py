from openpyxl import load_workbook
from pathlib import Path

source = Path('/home/ubuntu/sangkha-inspection/migration/sheets_copy.xlsx')
output = Path('/home/ubuntu/sangkha-inspection/migration/sheets_schema.txt')
workbook = load_workbook(source, read_only=True, data_only=True)
lines = []
for sheet in workbook.worksheets:
    rows = sheet.iter_rows(min_row=1, max_row=3, values_only=True)
    sample = list(rows)
    headers = [str(value).strip() if value is not None else '' for value in (sample[0] if sample else ())]
    non_empty_headers = [header for header in headers if header]
    lines.append(f'[{sheet.title}]')
    lines.append('HEADERS: ' + ' | '.join(non_empty_headers))
    if len(sample) > 1:
        lines.append('COLUMN_COUNT: ' + str(len(headers)))
    lines.append('')
output.write_text('\n'.join(lines), encoding='utf-8')
print(output)
