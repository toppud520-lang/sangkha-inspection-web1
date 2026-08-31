from openpyxl import load_workbook

wb = load_workbook('/tmp/real_data.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
for n, row in enumerate(ws.iter_rows(max_row=260, values_only=True), start=1):
    values = list(row[:8])
    if any(value not in (None, '') for value in values):
        print(n, values)
